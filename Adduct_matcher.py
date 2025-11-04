#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ESI Adduct Matcher
Automatically identify adduct pairs in mass spectrometry data
"""

import pandas as pd
import numpy as np
from pathlib import Path
from typing import Optional, List, Dict, Tuple
import warnings
warnings.filterwarnings('ignore')


class Adduct_matcher:
    """ESI Adduct Matcher"""
    
    def __init__(self, ppm_tolerance: float = 20.0, custom_adduct_file: Optional[str] = None):
        """
        Initialize Adduct Matcher
        
        Parameters:
        -----------
        ppm_tolerance : float
            Mass error tolerance (ppm), default is 20 ppm
        custom_adduct_file : str, optional
            Path to custom adduct table Excel file
        """
        self.ppm_tolerance = ppm_tolerance / 1_000_000  # Convert to ratio
        self.adduct_table = self._load_adduct_table(custom_adduct_file)
        
        # Column information
        self.rt_col = None
        self.mz_col = None
        self.intensity_col = None
        self.all_columns = None
        
    def _create_adduct_table(self) -> pd.DataFrame:
        """Create default table with 23 common ESI adducts mass differences"""
        data = {
            'From': ['[M+H]+'] * 23,
            'To': [
                '[M+Li]+', '[M+NH4]+', '[M+Na]+', '[M+K]+', '[M+H3O]+',
                '[M+H2O+H]+', '[M+MeOH+H]+', '[M+EtOH+H]+', '[M+IPA+H]+',
                '[M+ACN+H]+', '[M+DMSO+H]+', '[M+H2O+Na]+', '[M+MeOH+Na]+',
                '[M+EtOH+Na]+', '[M+IPA+Na]+', '[M+ACN+Na]+', '[M+DMSO+Na]+',
                '[M+HCOOH+H]+', '[M+CH3COOH+H]+', '[M+H2CO3+H]+',
                '[M+H2SO4+H]+', '[M+Na+H]+', '[M+2Na-H]+'
            ],
            'Delta_Da': [
                6.00817839, 17.02654909, 21.98194424, 37.95588144, 18.01056467,
                18.01056468, 32.02621475, 46.04186481, 60.05751488, 41.0265491,
                78.01393599, 39.99250892, 54.00815898, 68.02380905, 82.03945911,
                63.00849334, 99.99588022, 46.0054793, 60.02112937, 62.00039392,
                97.96737972, 22.9892207, 43.96388847
            ]
        }
        return pd.DataFrame(data)
    
    def _load_adduct_table(self, custom_file: Optional[str] = None) -> pd.DataFrame:
        """
        Load adduct table from custom file or use default
        
        Parameters:
        -----------
        custom_file : str, optional
            Path to custom adduct table Excel file
            Expected columns: 'From', 'To', 'Delta_Da'
            
        Returns:
        --------
        pd.DataFrame
            Adduct table
        """
        if custom_file:
            try:
                # Load custom adduct table
                custom_df = pd.read_excel(custom_file)
                
                # Validate required columns
                required_cols = ['From', 'To', 'Delta_Da']
                if not all(col in custom_df.columns for col in required_cols):
                    print(f"⚠ Warning: Custom adduct table missing required columns: {required_cols}")
                    print(f"  Using default adduct table instead.")
                    return self._create_adduct_table()
                
                # Filter valid rows
                custom_df = custom_df[required_cols].dropna()
                
                if len(custom_df) == 0:
                    print(f"⚠ Warning: Custom adduct table is empty.")
                    print(f"  Using default adduct table instead.")
                    return self._create_adduct_table()
                
                print(f"✓ Loaded custom adduct table: {len(custom_df)} adducts")
                return custom_df
                
            except Exception as e:
                print(f"⚠ Warning: Failed to load custom adduct table: {str(e)}")
                print(f"  Using default adduct table instead.")
                return self._create_adduct_table()
        else:
            # Use default adduct table
            return self._create_adduct_table()
    
    def load_data(self, file_path: str) -> pd.DataFrame:
        """
        Load data and automatically identify columns (supports Excel, CSV, TSV)
        
        Parameters:
        -----------
        file_path : str
            File path
            
        Returns:
        --------
        pd.DataFrame
            DataFrame with all columns
        """
        file_path = str(file_path)
        
        # Select reading method based on file extension
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path, encoding='utf-8-sig')
        elif file_path.endswith('.tsv') or file_path.endswith('.txt'):
            df = pd.read_csv(file_path, sep='\t', encoding='utf-8-sig')
        elif file_path.endswith(('.xlsx', '.xls', '.xlsm', '.xlsb')):
            df = pd.read_excel(file_path)
        else:
            raise ValueError(f"Unsupported file format. Supported: .xlsx, .xls, .xlsm, .xlsb, .csv, .tsv, .txt")
        
        # Auto-identify RT, m/z, Intensity columns
        rt_col = self._find_column(df.columns, [
            'rt', 'retention time', 'retention_time', 'retentiontime',
            'rt (min)', 'rt(min)', 'retention time (min)', 'time',
            'r.t.', 'r.t', '保留時間', '滯留時間'
        ])
        mz_col = self._find_column(df.columns, [
            'm/z', 'mz', 'm_z', 'mass', 'mass to charge',
            'precursor ion m/z', 'precursor m/z', 'precursormz',
            '質量', '質荷比'
        ])
        intensity_col = self._find_column(df.columns, [
            'intensity', 'int', 'intens', 'abundance', 'abund', 'height', 'area',
            'precursor ion intensity', 'precursor intensity', 'precursorintensity',
            '強度', '峰高', '峰面積'
        ])
        
        if not rt_col or not mz_col or not intensity_col:
            missing = []
            if not rt_col: missing.append("RT (Retention Time)")
            if not mz_col: missing.append("m/z (Mass-to-Charge)")
            if not intensity_col: missing.append("Intensity")
            
            available_cols = "\nAvailable columns: " + ", ".join(df.columns.tolist())
            raise ValueError(f"Cannot identify columns: {', '.join(missing)}\nPlease ensure headers contain these column names{available_cols}")
        
        # Mark main columns
        self.rt_col = rt_col
        self.mz_col = mz_col
        self.intensity_col = intensity_col
        self.all_columns = list(df.columns)
        
        print(f"✓ Successfully loaded file: {Path(file_path).name}")
        print(f"  Data shape: {df.shape[0]} rows × {df.shape[1]} columns")
        print(f"\nIdentified columns:")
        print(f"  RT (Retention Time):  {rt_col}")
        print(f"  m/z (Mass-to-Charge): {mz_col}")
        print(f"  Intensity:            {intensity_col}")
        print(f"  Other columns kept:   {len(self.all_columns) - 3}")
        
        # Remove invalid data (only check m/z and intensity > 0, RT can be 0)
        original_count = len(df)
        df = df[(df[mz_col] > 0) & (df[intensity_col] > 0)]
        df = df.dropna(subset=[rt_col, mz_col, intensity_col])
        
        removed_count = original_count - len(df)
        if removed_count > 0:
            print(f"\n⚠ Removed {removed_count} invalid rows (m/z≤0, Intensity≤0, or missing values)")
        
        print(f"✓ Valid data: {len(df)} rows\n")
        
        return df.reset_index(drop=True)
    
    def _find_column(self, columns: List[str], possible_names: List[str]) -> Optional[str]:
        """
        Find matching column name
        
        Parameters:
        -----------
        columns : list
            All column names
        possible_names : list
            List of possible column names
            
        Returns:
        --------
        str or None
            Found column name
        """
        for col in columns:
            col_lower = str(col).lower().strip()
            for name in possible_names:
                if name in col_lower:
                    return col
        return None
    
    def match_adducts(self, df: pd.DataFrame, rt_tolerance: float = 0.5) -> pd.DataFrame:
        """
        Match adducts, keep all original columns
        When multiple bases compete for the same pair, keep all matches (list all possibilities)
        
        Parameters:
        -----------
        df : pd.DataFrame
            DataFrame with mass spectrometry data
        rt_tolerance : float
            RT tolerance (minutes), default is 0.5 minutes
            
        Returns:
        --------
        pd.DataFrame
            DataFrame with adduct matching results, keeping all original columns
        """
        print("Starting adduct matching...")
        
        # 確保數據按RT排序
        df = df.sort_values(by=self.rt_col).reset_index(drop=True)
        
        # 建立結果列表
        results = []
        
        # 對每個數據點進行比對
        for i, row in df.iterrows():
            current_rt = row[self.rt_col]
            current_mz = row[self.mz_col]
            current_intensity = row[self.intensity_col]
            
            # 找出相同或接近滯留時間的其他訊號
            rt_mask = abs(df[self.rt_col] - current_rt) <= rt_tolerance
            nearby_peaks = df[rt_mask & (df.index != i)]
            
            # 與附近的peak比對
            for _, nearby_peak in nearby_peaks.iterrows():
                nearby_rt = nearby_peak[self.rt_col]
                nearby_mz = nearby_peak[self.mz_col]
                nearby_intensity = nearby_peak[self.intensity_col]
                
                # 計算RT差異（用於後續篩選）
                rt_diff = abs(current_rt - nearby_rt)
                
                # 計算質量差 (大減小)
                if nearby_mz > current_mz:
                    mz_diff = nearby_mz - current_mz
                    base_mz = current_mz
                    base_rt = current_rt
                    base_row = row
                    pair_mz = nearby_mz
                    pair_rt = nearby_rt
                    pair_row = nearby_peak
                else:
                    mz_diff = current_mz - nearby_mz
                    base_mz = nearby_mz
                    base_rt = nearby_rt
                    base_row = nearby_peak
                    pair_mz = current_mz
                    pair_rt = current_rt
                    pair_row = row
                
                # 與加合物表比對
                for _, adduct in self.adduct_table.iterrows():
                    theoretical_delta = adduct['Delta_Da']
                    
                    # 計算ppm容許範圍 (使用較大的m/z作為參考，與VBA相同)
                    reference_mz = max(current_mz, nearby_mz)  # 使用較大的m/z
                    ppm_tolerance_da = self.ppm_tolerance * reference_mz  # 轉換為Da
                    
                    # 檢查質量差是否在容許範圍內
                    if abs(mz_diff - theoretical_delta) <= ppm_tolerance_da:
                        # 計算實際的PPM誤差（用於報告）
                        ppm_error_value = abs(mz_diff - theoretical_delta) / reference_mz * 1_000_000
                        
                        # 建立結果字典，包含所有原始欄位
                        result = {}
                        
                        # 添加Base化合物的所有欄位（加上Base_前綴）
                        for col in self.all_columns:
                            result[f'Base_{col}'] = base_row[col]
                        
                        # 添加Pair化合物的所有欄位（加上Pair_前綴）
                        for col in self.all_columns:
                            result[f'Pair_{col}'] = pair_row[col]
                        
                        # 添加比對資訊
                        result['Base_Adduct'] = '[M+H]+'
                        result['Pair_Adduct'] = adduct['To']
                        result['Theoretical_Delta_Da'] = theoretical_delta
                        result['Observed_Delta_Da'] = mz_diff
                        result['PPM_Error'] = round(ppm_error_value, 2)
                        result['RT_Diff'] = rt_diff  # 添加RT差異用於篩選
                        result['Reference_mz'] = reference_mz
                        result['Annotation'] = f"{adduct['To']} of Base (m/z {base_mz:.4f})"
                        
                        results.append(result)
        
        if results:
            results_df = pd.DataFrame(results)
            
            # 移除重複的配對（同一組Base-Pair-Adduct只保留一次）
            base_mz_col = f'Base_{self.mz_col}'
            base_rt_col = f'Base_{self.rt_col}'
            pair_mz_col = f'Pair_{self.mz_col}'
            pair_rt_col = f'Pair_{self.rt_col}'
            
            # 先去除完全相同的配對
            results_df = results_df.drop_duplicates(
                subset=[base_mz_col, base_rt_col, pair_mz_col, pair_rt_col, 'Pair_Adduct'], 
                keep='first'
            ).reset_index(drop=True)
            
            # 方案2: 保留所有配對，但在同一個Base有多個Pair時，選RT最近的
            # 當一個Base有多個Pair候選時，保留RT最接近的
            results_df = results_df.sort_values('RT_Diff')
            results_df = results_df.drop_duplicates(
                subset=[base_mz_col, base_rt_col, 'Pair_Adduct'],
                keep='first'
            ).reset_index(drop=True)
            
            # 當一個Pair有多個Base時，保留所有配對（不去重）
            # 這樣可以列出所有可能性
            
            # 移除RT_Diff欄位（用戶不需要看到）
            results_df = results_df.drop(columns=['RT_Diff'])
            
            print(f"✓ 找到 {len(results_df)} 個加合物配對")
            
            # 顯示加合物類型統計
            adduct_counts = results_df['Pair_Adduct'].value_counts()
            print(f"\n加合物類型分布:")
            for adduct, count in adduct_counts.head(5).items():
                print(f"  {adduct}: {count} 個")
            
            return results_df
        else:
            print("⚠ 未找到符合的加合物配對")
            return pd.DataFrame()
    
    def save_results(self, df: pd.DataFrame, original_file: str, 
                    results: pd.DataFrame, output_file: Optional[str] = None):
        """
        Save results to Excel file, marking matching results on original data
        
        Parameters:
        -----------
        df : pd.DataFrame
            Original data
        original_file : str
            Original file path
        results : pd.DataFrame
            Matching results
        output_file : str, optional
            Output file name, auto-generated if not specified
        """
        if output_file is None:
            # Auto-generate output file name
            file_path = Path(original_file)
            output_file = str(file_path.parent / f"{file_path.stem}_adduct_results.xlsx")
        
        print(f"\nSaving results to: {Path(output_file).name}")
        
        # 準備標記資訊
        df_marked = df.copy()
        
        # 新增識別欄位
        df_marked['Adduct_Type'] = '[M+H]+'  # 預設都是[M+H]+
        df_marked['Pair_mz'] = ''            # 配對的m/z值
        df_marked['PPM_Error'] = ''          # PPM誤差
        df_marked['Description'] = ''     # 配對說明
        df_marked['Is_Matched_Base'] = False  # 是否為有配對的Base
        
        if not results.empty:
            # 建立m/z到索引的映射
            mz_col = self.mz_col
            rt_col = self.rt_col
            base_mz_col = f'Base_{mz_col}'
            pair_mz_col = f'Pair_{mz_col}'
            base_rt_col = f'Base_{rt_col}'
            pair_rt_col = f'Pair_{rt_col}'
            
            # 建立Pair到多個Base的映射字典
            pair_to_bases = {}  # key: (pair_mz, pair_rt), value: list of (base_mz, adduct, ppm)
            
            for _, result_row in results.iterrows():
                base_mz = result_row[base_mz_col]
                base_rt = result_row[base_rt_col]
                pair_mz = result_row[pair_mz_col]
                pair_rt = result_row[pair_rt_col]
                adduct = result_row['Pair_Adduct']
                ppm = result_row['PPM_Error']
                
                pair_key = (pair_mz, pair_rt)
                if pair_key not in pair_to_bases:
                    pair_to_bases[pair_key] = []
                
                pair_to_bases[pair_key].append({
                    'base_mz': base_mz,
                    'adduct': adduct,
                    'ppm': ppm
                })
            
            # 標記Base化合物（有配對的[M+H]+）
            for _, result_row in results.iterrows():
                base_mz = result_row[base_mz_col]
                base_rt = result_row[base_rt_col]
                
                # 同時檢查 m/z 和 RT
                mask = (abs(df_marked[mz_col] - base_mz) < 0.0001) & \
                       (abs(df_marked[rt_col] - base_rt) < 0.0001)
                if mask.any():
                    df_marked.loc[mask, 'Is_Matched_Base'] = True
            
            # 標記Pair化合物（標記為對應的加合物類型，可能有多個）
            for pair_key, base_list in pair_to_bases.items():
                pair_mz, pair_rt = pair_key
                
                # 找到對應的Pair行
                mask = (abs(df_marked[mz_col] - pair_mz) < 0.0001) & \
                       (abs(df_marked[rt_col] - pair_rt) < 0.0001)
                
                if mask.any():
                    # 如果有多個Base，按PPM排序（最小的最可能）
                    base_list_sorted = sorted(base_list, key=lambda x: x['ppm'])
                    
                    # 合併所有加合物類型
                    adduct_types = '; '.join([b['adduct'] for b in base_list_sorted])
                    
                    # 合併所有配對說明（不包含PPM資訊）
                    descriptions = '; '.join([
                        f"{b['adduct']} of Base m/z {b['base_mz']:.4f}"
                        for b in base_list_sorted
                    ])
                    
                    # 合併所有可能的 Base m/z（用分號分隔）
                    pair_mz_list = '; '.join([f"{b['base_mz']:.4f}" for b in base_list_sorted])
                    
                    # 合併所有PPM誤差（用分號分隔）
                    ppm_list = '; '.join([f"{b['ppm']:.2f}" for b in base_list_sorted])
                    
                    df_marked.loc[mask, 'Adduct_Type'] = adduct_types
                    df_marked.loc[mask, 'Pair_mz'] = pair_mz_list  # 列出所有Base m/z
                    df_marked.loc[mask, 'PPM_Error'] = ppm_list  # 列出所有PPM誤差
                    df_marked.loc[mask, 'Description'] = descriptions
        
        # 寫入Excel並設定格式
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Sheet 1: 寫入標記後的原始數據（不包含Is_Matched_Base欄位）
            df_output = df_marked.drop(columns=['Is_Matched_Base'])
            df_output.to_excel(writer, sheet_name='All_Feature_Annotated', index=False)
            
            # Sheet 2: 只保留非加合物的訊號（白色背景：黑色和紅色字體）
            df_non_adduct = df_marked[df_marked['Adduct_Type'] == '[M+H]+'].copy()
            df_non_adduct = df_non_adduct.drop(columns=['Is_Matched_Base'])
            df_non_adduct.to_excel(writer, sheet_name='Non_Adduct_Feature', index=False)
            
            # 取得workbook和worksheet以設定格式
            from openpyxl.styles import PatternFill, Font
            workbook = writer.book
            
            # 格式化 Sheet 1: All_Feature_Annotated
            worksheet1 = writer.sheets['All_Feature_Annotated']
            
            if not results.empty:
                # 設定顏色
                yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                red_font = Font(color='FF0000', bold=True)
                
                # 遍歷每一行設定格式
                for row_idx in range(2, len(df_marked) + 2):
                    adduct_type_value = df_marked.iloc[row_idx-2]['Adduct_Type']
                    is_matched_base = df_marked.iloc[row_idx-2]['Is_Matched_Base']
                    
                    if adduct_type_value == '[M+H]+' and is_matched_base:
                        # 有配對的Base化合物 - 紅色字體
                        for col_idx in range(1, len(df_output.columns) + 1):
                            cell = worksheet1.cell(row=row_idx, column=col_idx)
                            cell.font = red_font
                    elif adduct_type_value and adduct_type_value != '[M+H]+':
                        # Pair化合物 - 黃色背景
                        for col_idx in range(1, len(df_output.columns) + 1):
                            cell = worksheet1.cell(row=row_idx, column=col_idx)
                            cell.fill = yellow_fill
                
                # 格式化 Intensity 欄位為科學記號
                intensity_col_idx = list(df_output.columns).index(self.intensity_col) + 1
                for row in range(2, len(df_marked) + 2):
                    cell = worksheet1.cell(row=row, column=intensity_col_idx)
                    cell.number_format = '0.00E+00'
                
                # 格式化 PPM_Error 欄位
                if 'PPM_Error' in df_output.columns:
                    ppm_col_idx = list(df_output.columns).index('PPM_Error') + 1
                    for row in range(2, len(df_marked) + 2):
                        cell = worksheet1.cell(row=row, column=ppm_col_idx)
                        if cell.value and cell.value != '':
                            cell.number_format = '0.00'
                
                # 調整欄寬 - Sheet 1
                for column in worksheet1.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet1.column_dimensions[column_letter].width = adjusted_width
            
            # 格式化 Sheet 2: Non_Adduct_Feature
            worksheet2 = writer.sheets['Non_Adduct_Feature']
            
            if not results.empty and len(df_non_adduct) > 0:
                red_font = Font(color='FF0000', bold=True)
                
                # 標記有配對的Base化合物（紅色字體）
                non_adduct_with_match = df_marked[
                    (df_marked['Adduct_Type'] == '[M+H]+') & 
                    (df_marked['Is_Matched_Base'] == True)
                ].index.tolist()
                
                # 建立索引映射
                non_adduct_indices = df_non_adduct.index.tolist()
                
                for row_idx in range(2, len(df_non_adduct) + 2):
                    original_idx = non_adduct_indices[row_idx - 2]
                    
                    if original_idx in non_adduct_with_match:
                        # 有配對的Base - 紅色字體
                        for col_idx in range(1, len(df_non_adduct.columns) + 1):
                            cell = worksheet2.cell(row=row_idx, column=col_idx)
                            cell.font = red_font
                
                # 格式化 Intensity 欄位
                if self.intensity_col in df_non_adduct.columns:
                    intensity_col_idx = list(df_non_adduct.columns).index(self.intensity_col) + 1
                    for row in range(2, len(df_non_adduct) + 2):
                        cell = worksheet2.cell(row=row, column=intensity_col_idx)
                        cell.number_format = '0.00E+00'
                
                # 調整欄寬 - Sheet 2
                for column in worksheet2.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet2.column_dimensions[column_letter].width = adjusted_width
        
        print(f"✓ Results saved successfully!")
        print(f"  Sheet 1: All_Feature_Annotated - All peaks with annotations")
        print(f"  Sheet 2: Non_Adduct_Feature - Only [M+H]+ peaks (no adduct peaks)")
        if not results.empty:
            base_count = df_marked['Is_Matched_Base'].sum()
            pair_count = (df_marked['Adduct_Type'] != '[M+H]+').sum()
            unmatched_count = len(df_marked) - base_count - pair_count
            non_adduct_total = len(df_non_adduct)
            
            print(f"\n  All_Feature_Annotated:")
            print(f"    • Black text = Unmatched [M+H]+ ({unmatched_count} peaks)")
            print(f"    • Red text = Matched Base compounds ([M+H]+) ({base_count} peaks)")
            print(f"    • Yellow background = Matched adducts ({pair_count} peaks)")
            print(f"\n  Non_Adduct_Feature:")
            print(f"    • Total [M+H]+ peaks: {non_adduct_total}")
            print(f"    • Black text = Unmatched ({unmatched_count} peaks)")
            print(f"    • Red text = Has adduct pair ({base_count} peaks)")
        else:
            print(f"    • No matching results found")
    
    def process(self, file_path: str, rt_tolerance: float = 0.5, 
                output_file: Optional[str] = None) -> pd.DataFrame:
        """
        Complete processing workflow
        
        Parameters:
        -----------
        file_path : str
            Input file path
        rt_tolerance : float
            RT tolerance (minutes)
        output_file : str, optional
            Output file name
            
        Returns:
        --------
        pd.DataFrame
            Matching results
        """
        print("="*70)
        print("ESI Adduct Matcher")
        print("="*70 + "\n")
        
        # Load data
        df = self.load_data(file_path)
        
        # Show data preview
        print("Data preview (first 5 rows):")
        print(df.head().to_string(index=False))
        print()
        
        # Execute matching
        results = self.match_adducts(df, rt_tolerance=rt_tolerance)
        
        # Save results
        if not results.empty:
            self.save_results(df, file_path, results, output_file)
            
            print("\n" + "="*70)
            print("✓ Analysis completed!")
            print("="*70)
            
            # Show key statistics
            print(f"\nFound {len(results)} adduct pairs")
            print(f"Average PPM error: {results['PPM_Error'].mean():.2f}")
            print(f"PPM error range: {results['PPM_Error'].min():.2f} - {results['PPM_Error'].max():.2f}")
        else:
            print("\n" + "="*70)
            print("⚠ No matching adduct pairs found")
            print("="*70)
            print("\nSuggestions:")
            print("  1. Increase RT tolerance (e.g., 0.1 or 0.2)")
            print("  2. Increase PPM tolerance (e.g., 30 or 50)")
            print("  3. Check data quality")
        
        return results


class Adduct_matcherGUI:
    """Graphical User Interface"""
    
    def __init__(self, root):
        import tkinter as tk
        from tkinter import filedialog, messagebox, ttk
        
        self.tk = tk
        self.ttk = ttk
        self.filedialog = filedialog
        self.messagebox = messagebox
        
        self.root = root
        self.root.title("ESI Adduct Matcher")
        self.root.geometry("700x750")
        self.root.configure(bg='#f0f4f8')
        
        self.input_file = None
        self.adduct_file = None
        
        # 現代配色方案
        self.colors = {
            'primary': '#2196F3',      # 藍色
            'secondary': '#4CAF50',    # 綠色
            'accent': '#FF9800',       # 橙色
            'danger': '#F44336',       # 紅色
            'bg_light': '#f0f4f8',     # 淺灰藍
            'bg_white': '#ffffff',     # 白色
            'text_dark': '#2c3e50',    # 深灰
            'text_light': '#7f8c8d',   # 淺灰
            'border': '#e0e6ed'        # 邊框灰
        }
        
        self.create_widgets()
    
    def create_widgets(self):
        tk = self.tk
        ttk = self.ttk
        
        # 設定 ttk 樣式
        style = ttk.Style()
        style.theme_use('clam')
        
        # 標題區域
        title_frame = tk.Frame(self.root, bg=self.colors['primary'], height=80)
        title_frame.pack(fill="x", pady=0)
        title_frame.pack_propagate(False)
        
        title_label = tk.Label(
            title_frame,
            text="🔬 ESI Adduct Matcher",
            font=("Segoe UI", 20, "bold"),
            bg=self.colors['primary'],
            fg='white'
        )
        title_label.pack(expand=True)
        
        subtitle_label = tk.Label(
            title_frame,
            text="Identify adduct pairs in mass spectrometry data",
            font=("Segoe UI", 9),
            bg=self.colors['primary'],
            fg='white'
        )
        subtitle_label.pack()
        
        # 主要內容區域
        main_frame = tk.Frame(self.root, bg=self.colors['bg_light'])
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Step 1: 檔案選擇區域
        step1_frame = tk.LabelFrame(
            main_frame,
            text="  Step 1: Select Files  ",
            font=("Segoe UI", 11, "bold"),
            bg=self.colors['bg_white'],
            fg=self.colors['text_dark'],
            relief="flat",
            borderwidth=2,
            highlightbackground=self.colors['border'],
            highlightthickness=1
        )
        step1_frame.pack(fill="x", pady=(0, 15))
        
        # 內部容器
        step1_inner = tk.Frame(step1_frame, bg=self.colors['bg_white'])
        step1_inner.pack(fill="x", padx=15, pady=15)
        
        # 輸入檔案
        input_container = tk.Frame(step1_inner, bg=self.colors['bg_white'])
        input_container.pack(fill="x", pady=(0, 10))
        
        tk.Label(
            input_container,
            text="📁 Input Data File:",
            font=("Segoe UI", 10, "bold"),
            bg=self.colors['bg_white'],
            fg=self.colors['text_dark']
        ).pack(anchor="w", pady=(0, 5))
        
        input_row = tk.Frame(input_container, bg=self.colors['bg_white'])
        input_row.pack(fill="x")
        
        self.file_label = tk.Label(
            input_row,
            text="No file selected",
            font=("Segoe UI", 9),
            bg=self.colors['bg_light'],
            fg=self.colors['text_light'],
            anchor="w",
            padx=10,
            pady=8,
            relief="flat"
        )
        self.file_label.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        tk.Button(
            input_row,
            text="Browse",
            command=self.select_file,
            bg=self.colors['primary'],
            fg='white',
            font=("Segoe UI", 9, "bold"),
            relief="flat",
            padx=20,
            pady=8,
            cursor="hand2",
            activebackground='#1976D2',
            activeforeground='white'
        ).pack(side="right")
        
        # 加合物表檔案
        adduct_container = tk.Frame(step1_inner, bg=self.colors['bg_white'])
        adduct_container.pack(fill="x")
        
        tk.Label(
            adduct_container,
            text="📋 Adduct Table (Optional):",
            font=("Segoe UI", 10, "bold"),
            bg=self.colors['bg_white'],
            fg=self.colors['text_dark']
        ).pack(anchor="w", pady=(0, 5))
        
        adduct_row = tk.Frame(adduct_container, bg=self.colors['bg_white'])
        adduct_row.pack(fill="x")
        
        self.adduct_label = tk.Label(
            adduct_row,
            text="Using default (23 adducts)",
            font=("Segoe UI", 9),
            bg=self.colors['bg_light'],
            fg=self.colors['text_light'],
            anchor="w",
            padx=10,
            pady=8,
            relief="flat"
        )
        self.adduct_label.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        tk.Button(
            adduct_row,
            text="Browse",
            command=self.select_adduct_file,
            bg=self.colors['accent'],
            fg='white',
            font=("Segoe UI", 9, "bold"),
            relief="flat",
            padx=20,
            pady=8,
            cursor="hand2",
            activebackground='#F57C00',
            activeforeground='white'
        ).pack(side="right")
        
        # Step 2: 參數設定區域
        step2_frame = tk.LabelFrame(
            main_frame,
            text="  Step 2: Parameter Settings  ",
            font=("Segoe UI", 11, "bold"),
            bg=self.colors['bg_white'],
            fg=self.colors['text_dark'],
            relief="flat",
            borderwidth=2,
            highlightbackground=self.colors['border'],
            highlightthickness=1
        )
        step2_frame.pack(fill="x", pady=(0, 15))
        
        step2_inner = tk.Frame(step2_frame, bg=self.colors['bg_white'])
        step2_inner.pack(fill="x", padx=15, pady=15)
        
        # 參數網格
        param_grid = tk.Frame(step2_inner, bg=self.colors['bg_white'])
        param_grid.pack(fill="x")
        
        # PPM容差
        ppm_frame = tk.Frame(param_grid, bg=self.colors['bg_white'])
        ppm_frame.pack(side="left", expand=True, fill="x", padx=(0, 10))
        
        tk.Label(
            ppm_frame,
            text="⚙️ PPM Tolerance:",
            font=("Segoe UI", 10, "bold"),
            bg=self.colors['bg_white'],
            fg=self.colors['text_dark']
        ).pack(anchor="w", pady=(0, 5))
        
        self.ppm_tolerance_var = tk.StringVar(value="20")
        ppm_entry = tk.Entry(
            ppm_frame,
            textvariable=self.ppm_tolerance_var,
            font=("Segoe UI", 10),
            bg=self.colors['bg_light'],
            fg=self.colors['text_dark'],
            relief="flat",
            justify="center"
        )
        ppm_entry.pack(fill="x", ipady=8)
        
        # RT容差
        rt_frame = tk.Frame(param_grid, bg=self.colors['bg_white'])
        rt_frame.pack(side="left", expand=True, fill="x")
        
        tk.Label(
            rt_frame,
            text="⏱️ RT Tolerance (min):",
            font=("Segoe UI", 10, "bold"),
            bg=self.colors['bg_white'],
            fg=self.colors['text_dark']
        ).pack(anchor="w", pady=(0, 5))
        
        self.rt_tolerance_var = tk.StringVar(value="0.5")
        rt_entry = tk.Entry(
            rt_frame,
            textvariable=self.rt_tolerance_var,
            font=("Segoe UI", 10),
            bg=self.colors['bg_light'],
            fg=self.colors['text_dark'],
            relief="flat",
            justify="center"
        )
        rt_entry.pack(fill="x", ipady=8)
        
        # Step 3: 執行按鈕
        step3_frame = tk.Frame(main_frame, bg=self.colors['bg_light'])
        step3_frame.pack(fill="x", pady=(0, 15))
        
        self.run_button = tk.Button(
            step3_frame,
            text="▶️  Start Matching",
            command=self.process_data,
            bg=self.colors['secondary'],
            fg='white',
            font=("Segoe UI", 12, "bold"),
            relief="flat",
            padx=40,
            pady=15,
            cursor="hand2",
            activebackground='#45a049',
            activeforeground='white'
        )
        self.run_button.pack(expand=True)
        
        # 狀態顯示區域
        status_frame = tk.LabelFrame(
            main_frame,
            text="  Status & Results  ",
            font=("Segoe UI", 11, "bold"),
            bg=self.colors['bg_white'],
            fg=self.colors['text_dark'],
            relief="flat",
            borderwidth=2,
            highlightbackground=self.colors['border'],
            highlightthickness=1
        )
        status_frame.pack(fill="both", expand=True)
        
        status_inner = tk.Frame(status_frame, bg=self.colors['bg_white'])
        status_inner.pack(fill="both", expand=True, padx=10, pady=10)
        
        # 使用Text widget顯示狀態
        self.status_text = tk.Text(
            status_inner,
            height=12,
            font=("Consolas", 9),
            bg='#f8f9fa',
            fg=self.colors['text_dark'],
            relief="flat",
            wrap="word",
            state="disabled",
            padx=10,
            pady=10
        )
        self.status_text.pack(fill="both", expand=True)
        
        # 初始訊息
        self.update_status("👋 Welcome! Please select your data file to begin.")
    
    def select_file(self):
        """Select input file"""
        file_path = self.filedialog.askopenfilename(
            title="Select mass spectrometry data file",
            filetypes=[
                ("All supported formats", "*.xlsx *.xls *.xlsm *.xlsb *.csv *.tsv *.txt"),
                ("Excel files", "*.xlsx *.xls *.xlsm *.xlsb"),
                ("CSV files", "*.csv"),
                ("TSV files", "*.tsv *.txt"),
                ("All files", "*.*")
            ]
        )
        
        if file_path:
            self.input_file = file_path
            self.file_label.config(text=Path(file_path).name, fg="black")
    
    def select_adduct_file(self):
        """Select custom adduct table file"""
        file_path = self.filedialog.askopenfilename(
            title="Select custom adduct table (Excel)",
            filetypes=[
                ("Excel files", "*.xlsx *.xls *.xlsm *.xlsb"),
                ("All files", "*.*")
            ]
        )
        
        if file_path:
            self.adduct_file = file_path
            self.adduct_label.config(text=Path(file_path).name, fg="black")
    
    def update_status(self, message):
        """更新狀態顯示"""
        self.status_text.config(state="normal")
        self.status_text.insert("end", message + "\n")
        self.status_text.see("end")
        self.status_text.config(state="disabled")
        self.root.update()
    
    def process_data(self):
        """Process data"""
        if not self.input_file:
            self.messagebox.showerror("❌ Error", "Please select an input file first!")
            return
        
        try:
            # 清空狀態
            self.status_text.config(state="normal")
            self.status_text.delete(1.0, "end")
            self.status_text.config(state="disabled")
            
            # 禁用按鈕防止重複點擊
            self.run_button.config(state="disabled", bg='#cccccc')
            self.root.update()
            
            # 讀取參數
            ppm_tol = float(self.ppm_tolerance_var.get())
            rt_tol = float(self.rt_tolerance_var.get())
            
            self.update_status("="*60)
            self.update_status("🚀 Starting process...")
            self.update_status("="*60)
            
            # Create matcher with optional custom adduct table
            matcher = Adduct_matcher(ppm_tolerance=ppm_tol, custom_adduct_file=self.adduct_file)
            
            # 載入數據
            self.update_status("\n📂 Loading data...")
            df = matcher.load_data(self.input_file)
            
            # 顯示識別的欄位
            self.update_status(f"\n✅ Identified columns:")
            self.update_status(f"  • RT: {matcher.rt_col}")
            self.update_status(f"  • m/z: {matcher.mz_col}")
            self.update_status(f"  • Intensity: {matcher.intensity_col}")
            self.update_status(f"  • Other columns kept: {len(matcher.all_columns) - 3}")
            
            # 執行比對
            self.update_status("\n🔍 Executing adduct matching...")
            results = matcher.match_adducts(df, rt_tolerance=rt_tol)
            
            if not results.empty:
                # 顯示加合物類型統計
                adduct_counts = results['Pair_Adduct'].value_counts()
                self.update_status(f"\n📊 Adduct type distribution:")
                for adduct, count in adduct_counts.head(5).items():
                    self.update_status(f"  • {adduct}: {count} peaks")
                
                # 生成輸出檔名
                input_path = Path(self.input_file)
                output_path = input_path.parent / f"{input_path.stem}_adduct_results.xlsx"
                
                # 儲存結果
                self.update_status("\n💾 Saving results...")
                matcher.save_results(df, self.input_file, results, str(output_path))
                
                # 顯示統計
                self.update_status("\n" + "="*60)
                self.update_status("✅ Processing completed!")
                self.update_status("="*60)
                self.update_status(f"\n🎉 Found {len(results)} adduct pairs")
                self.update_status(f"📈 Average PPM error: {results['PPM_Error'].mean():.2f}")
                self.update_status(f"📊 PPM error range: {results['PPM_Error'].min():.2f} - {results['PPM_Error'].max():.2f}")
                self.update_status(f"\n💾 Results saved to:\n   {output_path}")
                
                self.messagebox.showinfo(
                    "✅ Completed", 
                    f"Processing completed successfully!\n\n"
                    f"📊 Found {len(results)} adduct pairs\n\n"
                    f"💾 Results saved to:\n{output_path.name}"
                )
            else:
                self.update_status("\n" + "="*60)
                self.update_status("⚠️ No matching adduct pairs found")
                self.update_status("="*60)
                self.update_status("\n💡 Suggestions:")
                self.update_status("  1️⃣ Increase RT tolerance (e.g., 1.0 or 2.0)")
                self.update_status("  2️⃣ Increase PPM tolerance (e.g., 30 or 50)")
                self.update_status("  3️⃣ Check data quality")
                
                self.messagebox.showwarning(
                    "⚠️ Notice", 
                    "No matching adduct pairs found\n\n"
                    "Please try adjusting parameters or check data quality"
                )
            
        except Exception as e:
            self.messagebox.showerror("❌ Error", f"Error during processing:\n\n{str(e)}")
            self.update_status(f"\n❌ Error: {str(e)}")
        finally:
            # 重新啟用按鈕
            self.run_button.config(state="normal", bg=self.colors['secondary'])


def main():
    """主程式"""
    import tkinter as tk
    from tkinter import filedialog, messagebox
    
    root = tk.Tk()
    app = Adduct_matcherGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
